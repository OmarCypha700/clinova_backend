from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from exams.models import Program, Student, Procedure, ProcedureStep


class Command(BaseCommand):
    help = 'Import all data from a single Excel file with multiple sheets'

    def add_arguments(self, parser):
        parser.add_argument(
            'filename',
            type=str,
            help='Excel file to import'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview import without saving to database'
        )

    def handle(self, *args, **options):
        filename = options['filename']
        dry_run = options['dry_run']
        
        if dry_run:
            self.stdout.write(self.style.WARNING(
                '=== DRY RUN MODE - No changes will be saved ===\n'
            ))
        
        try:
            wb = load_workbook(filename)
        except FileNotFoundError:
            raise CommandError(f'File not found: {filename}')
        except Exception as e:
            raise CommandError(f'Error loading file: {str(e)}')
        
        # Statistics
        stats = {
            'programs': {'created': 0, 'updated': 0, 'errors': 0},
            'students': {'created': 0, 'updated': 0, 'errors': 0},
            'procedures': {'created': 0, 'updated': 0, 'errors': 0},
            'steps': {'created': 0, 'updated': 0, 'errors': 0},
        }
        
        with transaction.atomic():
            # Import Programs
            if 'Programs' in wb.sheetnames:
                self.stdout.write('Importing programs...')
                ws = wb['Programs']
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not any(row):
                        continue
                    
                    try:
                        prog_id, name, abbreviation = row[0], row[1], row[2]
                        
                        if not name:
                            self.stdout.write(self.style.WARNING(
                                f'  Row {row_idx}: Skipped - missing name'
                            ))
                            stats['programs']['errors'] += 1
                            continue
                        
                        if prog_id:
                            # Update existing
                            program, created = Program.objects.update_or_create(
                                id=prog_id,
                                defaults={
                                    'name': name,
                                    'abbreviation': abbreviation or None
                                }
                            )
                        else:
                            # Create new
                            program, created = Program.objects.get_or_create(
                                name=name,
                                defaults={'abbreviation': abbreviation or None}
                            )
                        
                        if created:
                            stats['programs']['created'] += 1
                        else:
                            stats['programs']['updated'] += 1
                    
                    except Exception as e:
                        self.stdout.write(self.style.ERROR(
                            f'  Row {row_idx}: Error - {str(e)}'
                        ))
                        stats['programs']['errors'] += 1
                
                self.stdout.write(self.style.SUCCESS(
                    f'✓ Programs: {stats["programs"]["created"]} created, '
                    f'{stats["programs"]["updated"]} updated, '
                    f'{stats["programs"]["errors"]} errors'
                ))
            
            # Import Students
            if 'Students' in wb.sheetnames:
                self.stdout.write('Importing students...')
                ws = wb['Students']
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not any(row):
                        continue
                    
                    try:
                        student_id, index_number, full_name, program_name, is_active = row
                        
                        if not index_number or not full_name or not program_name:
                            self.stdout.write(self.style.WARNING(
                                f'  Row {row_idx}: Skipped - missing required fields'
                            ))
                            stats['students']['errors'] += 1
                            continue
                        
                        # Get program
                        try:
                            program = Program.objects.get(name=program_name)
                        except Program.DoesNotExist:
                            self.stdout.write(self.style.ERROR(
                                f'  Row {row_idx}: Program not found: {program_name}'
                            ))
                            stats['students']['errors'] += 1
                            continue
                        
                        # Parse is_active
                        active = is_active in ['Yes', 'True', '1', 1, True]
                        
                        # Create or update
                        student, created = Student.objects.update_or_create(
                            index_number=index_number,
                            defaults={
                                'full_name': full_name,
                                'program': program,
                                'is_active': active
                            }
                        )
                        
                        if created:
                            stats['students']['created'] += 1
                        else:
                            stats['students']['updated'] += 1
                    
                    except Exception as e:
                        self.stdout.write(self.style.ERROR(
                            f'  Row {row_idx}: Error - {str(e)}'
                        ))
                        stats['students']['errors'] += 1
                
                self.stdout.write(self.style.SUCCESS(
                    f'✓ Students: {stats["students"]["created"]} created, '
                    f'{stats["students"]["updated"]} updated, '
                    f'{stats["students"]["errors"]} errors'
                ))
            
            # Import Procedures
            if 'Procedures' in wb.sheetnames:
                self.stdout.write('Importing procedures...')
                ws = wb['Procedures']
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not any(row):
                        continue
                    
                    try:
                        proc_id, program_name, procedure_name, total_score = row
                        
                        if not program_name or not procedure_name:
                            self.stdout.write(self.style.WARNING(
                                f'  Row {row_idx}: Skipped - missing required fields'
                            ))
                            stats['procedures']['errors'] += 1
                            continue
                        
                        # Get program
                        try:
                            program = Program.objects.get(name=program_name)
                        except Program.DoesNotExist:
                            self.stdout.write(self.style.ERROR(
                                f'  Row {row_idx}: Program not found: {program_name}'
                            ))
                            stats['procedures']['errors'] += 1
                            continue
                        
                        # Create or update
                        procedure, created = Procedure.objects.update_or_create(
                            program=program,
                            name=procedure_name,
                            defaults={'total_score': total_score or 0}
                        )
                        
                        if created:
                            stats['procedures']['created'] += 1
                        else:
                            stats['procedures']['updated'] += 1
                    
                    except Exception as e:
                        self.stdout.write(self.style.ERROR(
                            f'  Row {row_idx}: Error - {str(e)}'
                        ))
                        stats['procedures']['errors'] += 1
                
                self.stdout.write(self.style.SUCCESS(
                    f'✓ Procedures: {stats["procedures"]["created"]} created, '
                    f'{stats["procedures"]["updated"]} updated, '
                    f'{stats["procedures"]["errors"]} errors'
                ))
            
            # Import Procedure Steps
            if 'Procedure Steps' in wb.sheetnames:
                self.stdout.write('Importing procedure steps...')
                ws = wb['Procedure Steps']
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not any(row):
                        continue
                    
                    try:
                        step_id, procedure_name, step_order, description = row
                        
                        if not procedure_name or not description:
                            self.stdout.write(self.style.WARNING(
                                f'  Row {row_idx}: Skipped - missing required fields'
                            ))
                            stats['steps']['errors'] += 1
                            continue
                        
                        # Get procedure
                        try:
                            procedure = Procedure.objects.get(name=procedure_name)
                        except Procedure.DoesNotExist:
                            self.stdout.write(self.style.ERROR(
                                f'  Row {row_idx}: Procedure not found: {procedure_name}'
                            ))
                            stats['steps']['errors'] += 1
                            continue
                        
                        # Create or update
                        step, created = ProcedureStep.objects.update_or_create(
                            procedure=procedure,
                            step_order=step_order or 1,
                            defaults={'description': description}
                        )
                        
                        if created:
                            stats['steps']['created'] += 1
                        else:
                            stats['steps']['updated'] += 1
                    
                    except Exception as e:
                        self.stdout.write(self.style.ERROR(
                            f'  Row {row_idx}: Error - {str(e)}'
                        ))
                        stats['steps']['errors'] += 1
                
                self.stdout.write(self.style.SUCCESS(
                    f'✓ Procedure Steps: {stats["steps"]["created"]} created, '
                    f'{stats["steps"]["updated"]} updated, '
                    f'{stats["steps"]["errors"]} errors'
                ))
            
            # Rollback if dry run
            if dry_run:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING(
                    '\n=== DRY RUN - All changes rolled back ==='
                ))
        
        # Summary
        self.stdout.write(self.style.SUCCESS('\n✓ Import completed'))
        total_created = sum(s['created'] for s in stats.values())
        total_updated = sum(s['updated'] for s in stats.values())
        total_errors = sum(s['errors'] for s in stats.values())
        
        self.stdout.write(f'  Total created: {total_created}')
        self.stdout.write(f'  Total updated: {total_updated}')
        if total_errors > 0:
            self.stdout.write(self.style.WARNING(f'  Total errors: {total_errors}'))