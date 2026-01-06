from django.core.management.base import BaseCommand
from django.utils import timezone
from openpyxl import Workbook
from exams.models import Program, Student, Procedure, ProcedureStep
import os


class Command(BaseCommand):
    help = 'Export all data to a single Excel file with multiple sheets'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output',
            type=str,
            default='complete_export.xlsx',
            help='Output filename'
        )

    def handle(self, *args, **options):
        output_file = options['output']
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Export Programs
        self.stdout.write('Exporting programs...')
        ws_programs = wb.create_sheet('Programs')
        ws_programs.append(['ID', 'Name', 'Abbreviation'])
        
        for program in Program.objects.all():
            ws_programs.append([
                program.id,
                program.name,
                program.abbreviation or ''
            ])
        
        self.stdout.write(self.style.SUCCESS(f'✓ {Program.objects.count()} programs exported'))
        
        # Export Students
        self.stdout.write('Exporting students...')
        ws_students = wb.create_sheet('Students')
        ws_students.append([
            'ID', 'Index Number', 'Full Name', 'Program Name', 'Is Active'
        ])
        
        for student in Student.objects.select_related('program').all():
            ws_students.append([
                student.id,
                student.index_number,
                student.full_name,
                student.program.name,
                'Yes' if student.is_active else 'No'
            ])
        
        self.stdout.write(self.style.SUCCESS(f'✓ {Student.objects.count()} students exported'))
        
        # Export Procedures
        self.stdout.write('Exporting procedures...')
        ws_procedures = wb.create_sheet('Procedures')
        ws_procedures.append([
            'ID', 'Program Name', 'Procedure Name', 'Total Score'
        ])
        
        for procedure in Procedure.objects.select_related('program').all():
            ws_procedures.append([
                procedure.id,
                procedure.program.name,
                procedure.name,
                procedure.total_score
            ])
        
        self.stdout.write(self.style.SUCCESS(f'✓ {Procedure.objects.count()} procedures exported'))
        
        # Export Procedure Steps
        self.stdout.write('Exporting procedure steps...')
        ws_steps = wb.create_sheet('Procedure Steps')
        ws_steps.append([
            'ID', 'Procedure Name', 'Step Order', 'Description'
        ])
        
        for step in ProcedureStep.objects.select_related('procedure').all():
            ws_steps.append([
                step.id,
                step.procedure.name,
                step.step_order,
                step.description
            ])
        
        self.stdout.write(self.style.SUCCESS(f'✓ {ProcedureStep.objects.count()} procedure steps exported'))
        
        # Add metadata sheet
        ws_meta = wb.create_sheet('Metadata', 0)
        ws_meta.append(['Export Information'])
        ws_meta.append(['Export Date', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws_meta.append(['Programs', Program.objects.count()])
        ws_meta.append(['Students', Student.objects.count()])
        ws_meta.append(['Procedures', Procedure.objects.count()])
        ws_meta.append(['Procedure Steps', ProcedureStep.objects.count()])
        ws_meta.append([])
        ws_meta.append(['Import Instructions'])
        ws_meta.append(['1. Keep sheet names unchanged'])
        ws_meta.append(['2. Import in order: Programs → Students → Procedures → Steps'])
        ws_meta.append(['3. For new records, leave ID column empty'])
        ws_meta.append(['4. For updates, keep the ID value'])
        
        # Save file
        wb.save(output_file)
        
        self.stdout.write(self.style.SUCCESS(
            f'\n✓ Complete export saved to {output_file}'
        ))
        self.stdout.write(f'  File size: {os.path.getsize(output_file) / 1024:.2f} KB')